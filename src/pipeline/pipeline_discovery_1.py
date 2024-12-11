import base64
import datetime
import logging
import os
import sys
import re
import time
import requests
import json
import openpyxl
import pandas as pd
from requests.auth import HTTPBasicAuth
from openpyxl.styles import Font


# Add the 'src' directory to the system path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))
from src.pipeline.build_pipeline_db import db_post_build_pipeline , db_get_build_pipeline
from src.pipeline.release_pipeline_db import db_post_release_pipeline
from src.dbDetails.migration_details_db import db_get_migration_details


# Setup logging
id = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
log_file = f'src\pipeline\logs\logs_pipelines_{id}'

# Clear the log file before starting a new run
with open(log_file, 'w'):
    pass

logging.basicConfig(filename=log_file, 
                    level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

def log_print(msg):
    logging.info(msg)
    print(msg)

def get_pipeline_details(project_):
    try:
        count =0
        url = f"{instance}/{project_}/_apis/build/definitions?api-version=6.0"
        response = requests.get(url, auth=HTTPBasicAuth(username, pat))
        if response.status_code == 200:
            data = response.json()
            ws.title = "Pipeline Data"
           
            if data['count'] > 0:
                log_print(f"Total Build Pipeline Count - {data['count']}")
                for pipeline in data['value']:
                    count = count+1
                    pipeline_id = pipeline['id']
                    log_print(f"Processing the pipeline Id {pipeline_id}")
                    file_url = f"{instance}/{project_}/_apis/pipelines/{pipeline_id}?revision=1"
                    response = requests.get(file_url, auth=HTTPBasicAuth(username, pat))
                    data = response.json()
                    variable_groups ={}
                    variables={}
                    detailedJson = data['configuration'].get('designerJson')
                    agent_details = ""
                    artifact_details=""
                    phases_=[]
                    if detailedJson:
                        variables = data['configuration']['designerJson'].get('variables', "")
                        variables_count= len(variables)
                        variable_groups = data['configuration']['designerJson'].get('variableGroups',"")
                        variable_groups_count= len(variable_groups)
                    else:
                        variables = data['configuration'].get('variables', "")
                        variables_count= len(variables)
                        variable_groups = data['configuration'].get('variableGroups',"")
                        variable_groups_count= len(variable_groups)
                        

                    path = data['configuration'].get('path', 'NA')
                    agent_details=pipeline.get('queue', {}).get('name')

                    if path == 'NA':
                        is_classic = 'Yes'
                    else:
                        is_classic = 'No (Build)'

                    collection = instance.split('/')

                    url = f"{instance}/{project_}/_apis/build/definitions/{pipeline_id}?api-version=6.1-preview"
                    response = requests.get(url , auth=HTTPBasicAuth(username, pat))
                    if response.status_code == 200:
                        data = response.json()
                        if data['repository']:
                            repository_name = data['repository'].get('name')
                            repository_branch = data['repository'].get('defaultBranch')
                            repository_type = data['repository'].get('type')
                        phases_string =""
                        execution_type=""
                        maxConcurrency =0
                        continueOnError =''
                        if is_classic == 'Yes':
                            # agent_details = pipeline.get('queue', "").get('name')
                            phases = data['process'].get('phases',"")
                            if phases != {}:
                                for phase in phases:
                                    execution_option = phase['target'].get('executionOptions',"")
                                    if phase and execution_option!= "":
                                        execution_type = phase['target']['executionOptions'].get('type','')
                                        if execution_type != '' :
                                            maxConcurrency =phase['target']['executionOptions'].get('maxConcurrency', 0)
                                            continueOnError =phase['target']['executionOptions'].get('continueOnError', '')
                                        
                                        phases_.append(phase['name'])
                                        phases_string = ', '.join(phases_)
                        else:
                            response_text = fetch_pipeline_yaml(pipeline_id, project_)
                            result = False
                            is_script_available = pd.isna(response_text)
                            if is_script_available is False:
                                result =is_release_pipeline(response_text)
                            is_classic = 'No (Release)' if result else is_classic

                    else:
                        log_print(f"Failed to fetch additional details: {response.status_code}")
                        log_print("Response content: " + response.text)
                    

                    url=f"{instance}/{project_}/_apis/build/builds?definitions={pipeline_id}&api-version=6.0"
                    response = requests.get(url , auth=HTTPBasicAuth(username, pat))
                    if response.status_code == 200:
                        data = response.json()
                        build_count = data['count']
                        build = next((b for b in data.get('value', '') if b['status'] != 'notStarted' and b['result'] == 'succeeded'), None)
                        if build:
                            build_id = build['id']
                            url = f"{instance}/{project_}/_apis/build/builds/{build_id}/artifacts?api-version=6.0"
                            response = requests.get(url, auth=HTTPBasicAuth(username, pat))
                            if response.status_code == 200:
                                artifacts_data = response.json()
                                artifact_details = next((artifact['name'] for artifact in artifacts_data.get('value','')), None)        
                    path = path.replace("/","")
                    pipeline_info = [project_,
                        pipeline['id'],
                        pipeline['name'],
                        pipeline['createdDate'],
                        path,
                        variables_count,
                        variable_groups_count,
                        repository_type,
                        repository_name,
                        repository_branch,
                        is_classic,
                        agent_details,
                        phases_string.replace("'",''),
                        execution_type,
                        maxConcurrency,
                        continueOnError,
                        build_count,
                        artifact_details
                    ]
                    ws.append(pipeline_info)
                    data = {
                    "project_name":project_,
                        "pipeline_id":str(pipeline_id),
                        "pipeline_name":pipeline['name'],
                        "last_updated_date":pipeline['createdDate'],
                        "file_name":path,
                        "variables":variables_count,
                        "variable_groups":variable_groups_count,
                        "repository_type":repository_type,
                        "repository_name":repository_name,
                        "repository_branch":repository_branch,
                        "classic_pipeline":is_classic,
                        "agents":agent_details,
                        "phases":phases_string.replace("'",''),
                        "execution_type":execution_type,
                        "max_concurrency":maxConcurrency,
                        "continue_on_error": continueOnError,
                        "builds":build_count,
                        "artifacts":artifact_details
                    }
                    if pipeline_id == 13:
                        db_post_build_pipeline(data)
                    db_post_build_pipeline(data)
                    if count%10 == 0:
                        wb.save(excel_name)
                # df = pd.DataFrame(pipeline_data)
                # format_excel(df,"Builds")
                wb.save(excel_name)
            else:
                log_print(f"No pipeline found in this project - {project_}")
                wb.save(excel_name)
        else:
            log_print(f"Failed to retrieve data. Status Code: {response.status_code} - {response.text}")
    except Exception as e:
        wb.save(excel_name)
        log_print(f"Error Fetching the get_pipeline_details() - {e} ")


def is_release_pipeline(response_text):
    try:
        deploy_keywords = ['DeployStaging', 'DeployProduction','resources','DownloadBuildArtifacts']
        for keyword in deploy_keywords:
            if keyword.lower() in response_text.lower():
                return True  
        return False
    except Exception as e:
        wb.save(excel_name)
        log_print(f"Error Fetching the is_release_pipeline() - {e} ")


def fetch_pipeline_yaml(pipeline_id_discovery, project_):
    try:
        pipeline_id =pipeline_id_discovery
        pipeline_details_url = f"{instance}/{project_}/_apis/pipelines/{pipeline_id}?revision=1"
        pipeline_response = requests.get(pipeline_details_url,auth=HTTPBasicAuth(username, pat))
        if pipeline_response.status_code != 200:
            log_print(f"Error fetching pipeline details: {pipeline_response.status_code} - {pipeline_response.text}")
            return
        pipeline_data = pipeline_response.json()
        repo_id = pipeline_data["configuration"]["repository"]["id"]
        file_path = pipeline_data["configuration"]["path"]
        download_yaml_url = f"{instance}/{project_}/_apis/git/repositories/{repo_id}/items?path={file_path}&api-version=6.0"
        yaml_response = requests.get(
            download_yaml_url,
            auth=HTTPBasicAuth(username, pat)
        )
        if yaml_response.status_code != 200:
            log_print(f"Error downloading YAML file: {yaml_response.status_code} - {yaml_response.text}")
            return 
        return yaml_response.text
    except Exception as e:
        wb.save(excel_name)
        log_print(f"Error Fetching the fetch_pipeline_yaml() - {e} ")
        return 

def get_releases( project ):
    try:
        count =0
        url = f"{instance}/{project}/_apis/release/definitions?api-version=6.0"
        response = requests.get(url,auth=HTTPBasicAuth(username, pat))
        if response.status_code == 200:
            data = response.json()
            id = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")   # Format as "HH-MM-SS"
            excel_name = f"{folder_path}\discovery_release_{project}_{id}.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Release Data"
            headers = ['Project','Release ID', 'Name', 'Created Date', 'Last Updated Date', 'Variable', 'Variable Groups','No of Releases','Release Names', 'Artifacts', 'Agents', 'Parallel Execution Type', 'Max Agents', 'ContinueOnError', 'Concurrency Count', 'QueueDepth Count']
            ws.append(headers)
            # format_excel(ws)
            if data.get('count', 0) > 0:
                for release in data.get('value', []) :
                    count =count+1
                    release_id =release['id']
                    releases_name = release['name']
                    release_count=0
                    release_names=[]
                    release_names_string=""
                    log_print(f"Processing the release Id {release_id}")
                    url = f"{instance}/{project}/_apis/release/definitions/{release_id}?api-version=6.1-preview"
                    response = requests.get(url, auth=HTTPBasicAuth('', pat))
                    if response.status_code == 200:
                        release_pipeline = response.json()
                        variables = release_pipeline.get('variables')
                        variable_groups = release_pipeline.get('variableGroups')
                        artifacts_count = len(release_pipeline.get('artifacts'))
                        variable_groups_count = len(variable_groups)
                        variables_count = len(variables)
                        
                        environment_details = []                   
                        for environment in release_pipeline['environments']:
                            concurrencyCount = environment['executionPolicy'].get('concurrencyCount',"")
                            queueDepthCount = environment['executionPolicy'].get('queueDepthCount',"")                   
                            # if 'deployPhases' in environment:
                            for deploy_phase in environment.get('deployPhases',{}):                               
                                if 'deploymentInput' in deploy_phase:
                                    maxNumberOfAgents =''
                                    continueOnError =''
                                    parallel_execution_type = deploy_phase['deploymentInput']['parallelExecution'].get('parallelExecutionType', '')
                                    if parallel_execution_type != 'none':
                                        maxNumberOfAgents =deploy_phase['deploymentInput']['parallelExecution'].get('maxNumberOfAgents', '')
                                        continueOnError =deploy_phase['deploymentInput']['parallelExecution'].get('continueOnError', '')

                                    queue_id = deploy_phase['deploymentInput'].get('queueId',{})
                                    agent_pool_details = get_agent_pool_details(project, queue_id)
                                    if agent_pool_details:
                                        environment_details=agent_pool_details.get('name')
                    else:
                        log_print(f"Failed to fetch variables - {response.status_code}-{response.text}")
                    
                    url = f"{instance}/{project}/_apis/release/releases?api-version=6.0"
                    response=requests.get(url , auth=HTTPBasicAuth(username, pat))
                    if response.status_code == 200:
                        data= response.json()
                        if data['count'] > 0:
                            releases = data['value']
                            for release in releases:
                                def_name = release['releaseDefinition'].get('name')
                                if def_name == releases_name:
                                    release_count =release_count+1
                                    release_names.append(release['name'])
                                    release_names_string = ', '.join(release_names)


                    collection = instance.split('/')

                    releases_info = [project,
                        release['id'], 
                        release['name'],       
                        release['createdOn'], 
                        release['modifiedOn'],  
                        variables_count,
                        variable_groups_count,
                        release_count,
                        release_names_string,
                        artifacts_count,
                        environment_details,
                        parallel_execution_type,
                        maxNumberOfAgents,
                        continueOnError,
                        concurrencyCount,
                        queueDepthCount
                    ]
                    ws.append(releases_info)
                    data = {
                        "project_name":project,
                            "release_id"	: release['id'],
                            "release_name": release['name'],
                            "created_date" :release['createdOn'], 
                            "updated_date": release['modifiedOn'],  
                            "release_variable" :variables_count, 
                            "variable_groups"	:variable_groups_count,
                            "no_of_relaseses":release_count,
                            "release_names":  release_names_string,
                            "artifacts"	:artifacts_count,
                            "agents" :environment_details,
                            "parallel_execution_type" : parallel_execution_type,
                            "max_agents"	:  str(maxNumberOfAgents),
                            "continueon_error" : str(continueOnError),
                            "concurrency_count" : str(concurrencyCount),
                            "queuedepth_count" :str(queueDepthCount)
                    }
                    db_post_release_pipeline(data)
                    if count%10 ==0:
                        wb.save(excel_name)
                wb.save(excel_name)
            else:
                log_print(f"No release found in the project - {project}")  
                wb.save(excel_name)
        else:
            log_print(f"Failed to fetch releases. Status Code: {response.status_code} - {response.text}")
    except Exception as e:
        wb.save(excel_name)
        log_print(f"Error Fetching the get_releases() - {e} ")
    
def get_agent_pool_details(project, queue_id):
    try:
        url=f"{instance}/{project}/_apis/distributedtask/queues/{queue_id}?api-version=6.0-preview"
        response = requests.get(url, auth=HTTPBasicAuth(username, pat))
        if response.status_code == 200:
            data = response.json()
            return data
        else:
            return False
    except Exception as e:
        wb.save(excel_name)
        log_print(f"Error Fetching the get_agent_pool_details() - {e} ")


def format_excel(ws):
    try:
        # Make headers bold
        for cell in ws[1]:  # First row (headers), assuming openpyxl index starts from 1
            cell.font = Font(bold=True)

        # Set the column width dynamically based on content length
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Column index
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception as e:
                    pass
            adjusted_width = max_length + 2  # Add some padding
            ws.column_dimensions[ws.cell(row=1, column=column).column_letter].width = adjusted_width
    except Exception as e:
        print(f"Error in format_excel: {e}")


        # artifact_column_idx = df.columns.get_loc('Agents/Stages')
        # parallel_column_idx = df.columns.get_loc('Parallel Execution')
        # worksheet.set_column(artifact_column_idx, artifact_column_idx, 80, None, {'text_wrap': True})
        # worksheet.set_column(parallel_column_idx, parallel_column_idx, 50, None, {'text_wrap': True})
def create_mapping_sheet():
    results=db_get_build_pipeline()
    folder = f"src\pipeline"
    headers =["Source_Project","Source_Pipeline_Name","File_Name","Source_Repo_Name", "Source_Repo_Name","Target_Project", "Target_Pipeline_Name","Is_Classic", "Migration_Required","Status"]
    wb = openpyxl.Workbook()
    ws = wb.active
    # id = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")   # Format as "HH-MM-SS"
    excel_name =f"{folder}\mapping_migration.xlsx"
    ws.append(headers)
    for result in results:
        data =[
              result.project_name,
              result.pipeline_name,
              result.file_name,
              result.repository_name,
              result.repository_branch,
              result.project_name,
              result.pipeline_name,
              result.classic_pipeline,
              "yes",
              "Discovery Completed"
        ]
        ws.append(data)
    wb.save(excel_name)

if __name__ == "__main__":
    pipeline_data = []
    releases_data=[]
    
    folder_path=f"src\pipeline\output_folder"
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # # Read the Excel file
    # config_df = pd.read_excel(f'src\pipeline\input_discovery.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    headers =['Project','Pipeline ID', 'Name', 'Last Updated Date', 'FileName', 'Variables', 'Variable Groups', 'Repository Type','Repository Name' ,'Repository Branch', 'Classic Pipeline', 'Agents', 'Phases', 'Execution Type', 'Max Concurrency', 'ContinueOn Error', 'Builds', 'Artifacts']
            # format_excel(ws)
    ws.append(headers)

    results =db_get_migration_details()
    # Iterate over each row in the DataFrame
    for result in results:
        # Extract values from the current row
        instance = result.source_server_url
        project_name = result.source_project_name
        username = ""
        pat = result.source_pat

        log_print(f"--- {project_name} ---")
        excel_name =f"{folder_path}\source_discovery_build.xlsx"
        wb.save(excel_name)
        get_pipeline_details(project_name)
    wb.close()
    # for index, row in config_df.iterrows():
    #     wb = openpyxl.Workbook()
    #     ws = wb.active
    #     id = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")   # Format as "HH-MM-SS"
    #     excel_name =f"{folder_path}\discovery_release_{project_name}_{id}.xlsx"
    #     get_releases(project_name)
    #     wb.close()

    create_mapping_sheet()