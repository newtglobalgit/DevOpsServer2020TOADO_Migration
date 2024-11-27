import os
import pandas as pd
import requests
from requests.auth import HTTPBasicAuth
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import getpass
import logging
from urllib.parse import quote
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from collections import defaultdict
import gc
from utils.common import get_project_names, get_repo_names_by_project

log_dir = "logs"
if not os.path.exists(log_dir):
    os.makedirs(log_dir)
# Create a logger
logger = logging.getLogger()
# Set the log level
logger.setLevel(logging.INFO)
# File handler
file_handler = logging.FileHandler(os.path.join(log_dir, f'git_discovery_{datetime.now().strftime("%Y%m%d%H%M%S")}.log'))
file_handler.setLevel(logging.INFO)
# Console handler
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
# Log format
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)
# Add handlers to the logger
logger.addHandler(file_handler)
logger.addHandler(console_handler)


def encode_url_component(component):
    return quote(component, safe='')

    
def make_request_with_retries(url, pat, timeout=300, max_retries=10, backoff_factor=0.5):
    auth = HTTPBasicAuth('', pat)
    session = requests.Session()
    retries = Retry(
        total=max_retries,
        backoff_factor=backoff_factor,
        status_forcelist=[408, 429, 500, 502, 503, 504],
        allowed_methods=["HEAD", "GET", "OPTIONS"]
    )
    session.mount('https://', HTTPAdapter(max_retries=retries))
    for attempt in range(max_retries):
        try:
            response = session.get(url, auth=auth, timeout=timeout)
            if response.status_code == 200:
                logger.info(f"Request succeeded for URL {url}")
                return response
            elif response.status_code == 404:
                logger.info(f"Resource not found (404) for URL {url}.")
                return None
            elif response.status_code == 403:
                logger.error("Access forbidden (403). Check permissions or credentials.")
                return None
            else:
                logger.warning(f"Unexpected status code {response.status_code} for URL {url}.")
                return None
        except requests.exceptions.RequestException as e:
            logger.error(f"Error occurred while making request to {url}: {e}")
            return None
    logger.error(f"Exceeded max retries for URL {url}")
    return None


# Function to read configuration from Excel file
def read_config_from_excel(file_path):
    try:
        df = pd.read_excel(file_path, sheet_name='Input')
        print("DataFrame loaded successfully.")
        print("DataFrame content:")
        print(df)
    except ValueError as e:
        print(f"Error: {e}")
        return None, None, None, None, None, None

    server_url = df.at[0, 'Server URL'].strip()
    project = df.at[0, 'Project Name'].strip()
    pat = df.at[0, 'PAT'].strip()
    repository_name = df.at[0, 'Repository Name'].strip()
    branch_name = df.at[0, 'Branch Name'].strip() if not pd.isna(df.at[0, 'Branch Name']) else None

    if not server_url or not project or not pat or not repository_name:
        print("Please ensure the Excel file contains the required data.")
        return None, None, None, None, None, None

    return server_url, project, pat, repository_name, branch_name, df

    
def authenticate_and_get_projects(server_url, pat, api_version):
    url = f'{server_url}/_apis/projects?api-version={api_version}'
    response = make_request_with_retries(url, pat)
    if response:
        logger.info('Login successful.')
        return response.json()
    else:
        logger.error('Failed to authenticate.')
        return None

def get_repositories(server_url, project, pat, api_version, batch_size=50):
    """Retrieve all repositories with direct token handling, URL updates, and empty response checks."""
    encoded_project = encode_url_component(project)
    url = f'{server_url}/{encoded_project}/_apis/git/repositories?$top={batch_size}&api-version={api_version}'
    
    all_repositories = []
    repo_set = set()

    while True:
        response = make_request_with_retries(url, pat)
        
        if response:
            response_data = response.json()
            if 'value' in response_data and isinstance(response_data['value'], list):
                data = response_data['value']
                if not data:
                    logger.info("No repositories found in the current batch; stopping.")
                    break
                for repo in data:
                    if isinstance(repo, dict):  # Ensure repo is a dictionary
                        repo_id = repo.get('id')
                        if repo_id and repo_id not in repo_set:
                            all_repositories.append(repo)
                            repo_set.add(repo_id)
                    else:
                        logger.error(f"Unexpected format for repository item: {repo}")

                continuation_token = response.headers.get('x-ms-continuationtoken')
                if continuation_token:
                    url = f'{server_url}/{encoded_project}/_apis/git/repositories?$top={batch_size}&api-version={api_version}&continuationToken={continuation_token}'
                else:
                    break
            else:
                logger.error(f"Unexpected response format: {response_data}")
                break
        else:
            logger.error('Failed to retrieve repositories with pagination or no valid data found.')
            break

    return all_repositories


def get_branches(server_url, project, repository_id, pat, api_version, batch_size=50):
    encoded_project = encode_url_component(project)
    encoded_repository_id = encode_url_component(repository_id)
    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/refs?filter=heads&$top={batch_size}&api-version={api_version}'
    
    all_branches = []
    branch_set = set()

    while True:
        response = make_request_with_retries(url, pat)

        if response:
            response_data = response.json()
            data = response_data.get('value', [])
            if not data:
                break
            for branch in data:
                branch_name = branch.get('name')
                if branch_name and branch_name not in branch_set:
                    all_branches.append(branch)
                    branch_set.add(branch_name)

            continuation_token = response.headers.get('x-ms-continuationtoken')
            if continuation_token:
                url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/refs?filter=heads&$top={batch_size}&api-version={api_version}&continuationToken={continuation_token}'
            else:
                break
        else:
            break

    return all_branches



def get_latest_commit_info(server_url, project, repository_id, branch_names, pat, api_version):
    """Retrieve latest commit info for each branch in a batch."""
    latest_commits = {}
    for branch_name in branch_names:
        encoded_branch_name = encode_url_component(branch_name)
        url = f'{server_url}/{project}/_apis/git/repositories/{repository_id}/commits?searchCriteria.itemVersion.version={encoded_branch_name}&$top=1&api-version={api_version}'
        response = make_request_with_retries(url, pat)
        if response:
            commit = response.json().get('value', [None])[0]
            if commit:
                latest_commits[branch_name] = {
                    'commitId': commit['commitId'],
                    'comment': commit['comment'],
                    'author': commit['author']['name'],
                    'date': commit['author']['date']
                }
    return latest_commits


def get_files_in_branch(server_url, project, repository_id, branch_name, pat, api_version, batch_size=50):
    encoded_project = encode_url_component(project)
    encoded_repository_id = encode_url_component(repository_id)
    encoded_branch_name = encode_url_component(branch_name)
    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/items?scopePath=/&recursionLevel=Full&versionDescriptor[version]={encoded_branch_name}&$top={batch_size}&api-version={api_version}'
    
    all_files = []
    file_set = set()

    while True:
        response = make_request_with_retries(url, pat)

        if response:
            response_data = response.json()
            data = response_data.get('value', [])
            if not data:
                break
            for file_data in data:
                file_path = file_data.get('path')
                if file_path and file_path not in file_set:
                    all_files.append(file_data)
                    file_set.add(file_path)

            continuation_token = response.headers.get('x-ms-continuationtoken')
            if continuation_token:
                url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/items?scopePath=/&recursionLevel=Full&versionDescriptor[version]={encoded_branch_name}&$top={batch_size}&api-version={api_version}&continuationToken={continuation_token}'
            else:
                break
        else:
            break

    return all_files

def get_file_size(server_url, project, repository_id, sha1_list, pat, api_version):
    """Retrieve file sizes in batch using SHA1 list."""
    file_sizes = {}
    for sha1 in sha1_list:
        url = f'{server_url}/{project}/_apis/git/repositories/{repository_id}/blobs/{sha1}?api-version={api_version}'
        response = make_request_with_retries(url, pat)
        if response:
            file_sizes[sha1] = int(response.headers.get('Content-Length', 0))
    return file_sizes


def get_commit_count(server_url, project, repository_id, file_paths, pat, api_version):
    """Retrieve commit counts for each file in batch using file paths."""
    commit_counts = {}
    for file_path in file_paths:
        url = f'{server_url}/{project}/_apis/git/repositories/{repository_id}/commits?searchCriteria.itemPath={file_path}&api-version={api_version}'
        response = make_request_with_retries(url, pat)
        if response:
            commit_counts[file_path] = response.json().get('count', 1)
    return commit_counts

def get_all_commits(server_url, project, repository_id, branch_name, pat, api_version, batch_size=50):
    """Retrieve all commits in a branch with direct token handling, URL updates, and empty response checks."""
    encoded_project = encode_url_component(project)
    encoded_repository_id = encode_url_component(repository_id)
    encoded_branch_name = encode_url_component(branch_name)
    
    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/commits?searchCriteria.itemVersion.version={encoded_branch_name}&$top={batch_size}&api-version={api_version}'
    all_commits = []
    commit_set = set()

    while True:
        response = make_request_with_retries(url, pat)
        
        if response:
            response_data = response.json()
            # Ensure 'value' key exists and is a list
            if 'value' in response_data and isinstance(response_data['value'], list):
                data = response_data['value']
                if not data:
                    logger.info("No commits found in the current batch; stopping.")
                    break
                for commit in data:
                    # Check if each commit is a dictionary before accessing 'commitId'
                    if isinstance(commit, dict):
                        commit_id = commit.get('commitId')
                        if commit_id and commit_id not in commit_set:
                            all_commits.append(commit)
                            commit_set.add(commit_id)
                    else:
                        logger.error(f"Unexpected format for commit item: {commit}")

                continuation_token = response.headers.get('x-ms-continuationtoken')
                if continuation_token:
                    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/commits?searchCriteria.itemVersion.version={encoded_branch_name}&$top={batch_size}&api-version={api_version}&continuationToken={continuation_token}'
                else:
                    break
            else:
                logger.error(f"Unexpected response format: {response_data}")
                break
        else:
            logger.error('Failed to retrieve commits with pagination or no valid data found.')
            break

    return all_commits



def get_all_repo_commits(server_url, project, repository_id, pat, api_version, batch_size=50):
    """Retrieve all commits in a repository with direct token handling, URL updates, and empty response checks."""
    encoded_project = encode_url_component(project)
    encoded_repository_id = encode_url_component(repository_id)
    
    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/commits?$top={batch_size}&api-version={api_version}'
    all_commits = []
    commit_set = set()

    while True:
        response = make_request_with_retries(url, pat)
        
        if response:
            response_data = response.json()
            # Ensure 'value' key exists and is a list
            if 'value' in response_data and isinstance(response_data['value'], list):
                data = response_data['value']
                if not data:
                    logger.info("No repository commits found in the current batch; stopping.")
                    break
                for commit in data:
                    # Check if each commit is a dictionary before accessing 'commitId'
                    if isinstance(commit, dict):
                        commit_id = commit.get('commitId')
                        if commit_id and commit_id not in commit_set:
                            all_commits.append(commit)
                            commit_set.add(commit_id)
                    else:
                        logger.error(f"Unexpected format for commit item: {commit}")

                continuation_token = response.headers.get('x-ms-continuationtoken')
                if continuation_token:
                    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/commits?$top={batch_size}&api-version={api_version}&continuationToken={continuation_token}'
                else:
                    break
            else:
                logger.error(f"Unexpected response format: {response_data}")
                break
        else:
            logger.error('Failed to retrieve repository commits with pagination or no valid data found.')
            break

    return all_commits

def get_tags(server_url, project, repository_id, pat, api_version, batch_size=50):
    """Retrieve all tags in a repository with direct token handling, URL updates, and empty response checks."""
    encoded_project = encode_url_component(project)
    encoded_repository_id = encode_url_component(repository_id)
    
    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/refs?filter=tags&$top={batch_size}&api-version={api_version}'
    all_tags = []
    tag_set = set()

    while True:
        response = make_request_with_retries(url, pat)
        
        if response:
            response_data = response.json()
            # Ensure 'value' key exists and is a list
            if 'value' in response_data and isinstance(response_data['value'], list):
                data = response_data['value']
                if not data:
                    logger.info("No tags found in the current batch; stopping.")
                    break
                for tag in data:
                    # Check if each tag is a dictionary before accessing 'name'
                    if isinstance(tag, dict):
                        tag_name = tag.get('name')
                        if tag_name and tag_name not in tag_set:
                            all_tags.append(tag)
                            tag_set.add(tag_name)
                    else:
                        logger.error(f"Unexpected format for tag item: {tag}")

                continuation_token = response.headers.get('x-ms-continuationtoken')
                if continuation_token:
                    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/refs?filter=tags&$top={batch_size}&api-version={api_version}&continuationToken={continuation_token}'
                else:
                    break
            else:
                logger.error(f"Unexpected response format: {response_data}")
                break
        else:
            logger.error('Failed to retrieve tags with pagination or no valid data found.')
            break

    return all_tags


def get_tag_details(server_url, project, repository_id, tag_ids, pat, api_version):
    """Retrieve details for each tag in batch using tag IDs."""
    tag_details = {}
    for tag_id in tag_ids:
        url = f'{server_url}/{project}/_apis/git/repositories/{repository_id}/annotatedtags/{tag_id}?api-version=6.0-preview.1'
        response = make_request_with_retries(url, pat)
        if response:
            tag_details[tag_id] = response.json()
    return tag_details



def get_commit_details(server_url, project, repository_id, commit_id, pat, api_version):
    encoded_project = encode_url_component(project)
    encoded_repository_id = encode_url_component(repository_id)
    encoded_commit_id = encode_url_component(commit_id)
    
    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/commits/{encoded_commit_id}?api-version={api_version}'
    response = make_request_with_retries(url, pat)

    if response:
        return response.json()
    logger.error(f'Failed to retrieve commit details for {commit_id}. Skipping.')
    return None


def map_commit_tags(tags):
    commit_tag_map = {}
    for tag in tags:
        commit_id = tag['peeledObjectId'] if 'peeledObjectId' in tag else tag['objectId']
        commit_tag_map[commit_id] = tag['name']
    return commit_tag_map


def apply_header_styles(workbook, sheet_name):
    sheet = workbook[sheet_name]
    header_fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in sheet[1]:  # Assuming headers are in the first row
        cell.fill = header_fill
        cell.font = header_font


def apply_black_border(sheet):
    thin_border = Border(left=Side(style='thin', color='000000'),
                         right=Side(style='thin', color='000000'),
                         top=Side(style='thin', color='000000'),
                         bottom=Side(style='thin', color='000000'))
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border


def remove_gridlines(sheet):
    sheet.sheet_view.showGridLines = False


def adjust_column_width(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Get the column name
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length, 30)
        sheet.column_dimensions[column].width = adjusted_width


def align_cells(sheet):
    for row in sheet.iter_rows(min_row=2):  # Skip header row
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')
            elif isinstance(cell.value, datetime) or isinstance(cell.value, str) and '-' in cell.value and ':' in cell.value:
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')


def sanitize_sheet_name(name):
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    for char in invalid_chars:
        name = name.replace(char, '')
    return name[:31]  # Excel sheet names can have a maximum of 31 characters


def generate_report(data_source_code, data_commits, data_all_commits, data_tags, output_path, project_name, repo_name,server_url, start_time):
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Create an empty summary sheet
        pd.DataFrame().to_excel(writer, sheet_name='summary', index=False)
        df_source_code = pd.DataFrame(data_source_code)
        df_commits = pd.DataFrame(data_commits)
        df_all_commits = pd.DataFrame(data_all_commits)
        df_tags = pd.DataFrame(data_tags)
        df_source_code.to_excel(writer, sheet_name='source_code', index=False)
        df_commits.to_excel(writer, sheet_name='commits', index=False)
        df_tags.to_excel(writer, sheet_name='tags', index=False)
    
    workbook = load_workbook(output_path)
    worksheet_summary = workbook['summary']
    apply_header_styles(workbook, 'source_code')
    apply_header_styles(workbook, 'commits')
    apply_header_styles(workbook, 'tags')

    apply_black_border(workbook['summary'])
    apply_black_border(workbook['source_code'])
    apply_black_border(workbook['commits'])
    apply_black_border(workbook['tags'])

    remove_gridlines(workbook['summary'])
    remove_gridlines(workbook['source_code'])
    remove_gridlines(workbook['commits'])
    remove_gridlines(workbook['tags'])

    # Make header text bold for all sheets except the summary sheet
    for sheet_name in ['source_code', 'commits', 'tags']:
        sheet = workbook[sheet_name]
        apply_header_styles(workbook, sheet_name)
        apply_black_border(sheet)
        remove_gridlines(sheet)
        adjust_column_width(sheet)
        align_cells(sheet)

    run_duration = datetime.now() - start_time
    hours, remainder = divmod(run_duration.total_seconds(), 3600)
    minutes, seconds = divmod(remainder, 60)
    formatted_run_duration = f"{int(hours)} hours, {int(minutes)} minutes, {seconds:.1f} seconds"

    # Determine organization or collection based on the server URL
    if "dev.azure.com" in server_url:
        org_or_collection = "organization"
        org_or_collection_name = server_url.split("/")[-1]  # Extract organization name from URL
    else:
        org_or_collection = "collection"
        org_or_collection_name = server_url.split('/')[-1]  # Extract collection name from URL

    # Write summary data
    summary_data = {
        'Report Title': f"Project {project_name} Git Report.",
        'Purpose of the report': f"This provides a detailed view of the repo {repo_name} in project"
                                 f" {project_name} of {org_or_collection} {org_or_collection_name}.",
        'Run Date': datetime.now().strftime('%d-%b-%Y %I:%M %p'),
        'Run Duration': formatted_run_duration,
        'Run By': getpass.getuser(),
        'Input': f"Server URL: {server_url}, Project Name: {project_name}, Repository Name: {repo_name}"
    }

    row = 1
    for key, value in summary_data.items():
        worksheet_summary.cell(row=row, column=1).value = key
        worksheet_summary.cell(row=row, column=1).fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')
        worksheet_summary.cell(row=row, column=1).font = Font(color='FFFFFF', bold=True)
        worksheet_summary.cell(row=row, column=2).value = value
        worksheet_summary.cell(row=row, column=2).alignment = Alignment(horizontal='left')
        row += 1

    for i in range(len(summary_data)):
        worksheet_summary.column_dimensions['A'].width = 26.71  # Adjust column A width
        worksheet_summary.column_dimensions['B'].width = 85.87  # Adjust column B width
        worksheet_summary.row_dimensions[i + 1].height = 30  # Adjust row height

    apply_black_border(worksheet_summary)  # Apply black border to summary sheet

    workbook.save(output_path)

def process(server_url, pat, project, repository_name, branch_name, batch_size=50):
    api_version = '6.0'  # Adjust if your server uses a different version

    # Debug information
    print(f"Server URL: {server_url}")
    print(f"Project: {project}")
    print(f"Repo: {repository_name}")
    print(f"Branch: {branch_name}")

    repositories = get_repositories(server_url, project, pat, api_version, batch_size)
    if isinstance(repositories, list):  # Check that repositories is a list
        for repo in repositories:
            if isinstance(repo, dict) and repo.get('name') == repository_name:
                repo_id = repo['id']

                # Initialize data collections
                data_source_code = []
                data_commits = []
                data_all_commits = []
                data_tags = []

                # Get branches and retrieve latest commits in batch
                branches = get_branches(server_url, project, repo_id, pat, api_version, batch_size=batch_size)
                branch_names = [branch['name'].replace('refs/heads/', '') for branch in branches] if branch_name is None else [branch_name]

                # Fetch latest commits for each branch
                latest_commits = get_latest_commit_info(server_url, project, repo_id, branch_names, pat, api_version)

                for branch in branch_names:
                    commit_info = latest_commits.get(branch)
                    if commit_info:
                        commit_id = commit_info['commitId']
                        comment = commit_info['comment']
                        author = commit_info['author']
                        last_modified = commit_info['date']

                        # Retrieve all commits for the branch
                        commits = get_all_commits(server_url, project, repo_id, branch, pat, api_version, batch_size=batch_size)
                        for commit in commits:
                            commit_data = {
                                'Collection Name': server_url.split('/')[-1],
                                'Project Name': project,
                                'Repository Name': repository_name,
                                'Branch Name': branch,
                                'Commit ID': commit['commitId'],
                                'Commit Message': commit['comment'],
                                'Author': commit['author']['name'],
                                'Commit Date': commit['author']['date']
                            }
                            data_commits.append(commit_data)

                        # Get files in branch and retrieve file size and commit count in batch
                        files = get_files_in_branch(server_url, project, repo_id, branch, pat, api_version, batch_size=batch_size)
                        sha1_list = [file['objectId'] for file in files if not file.get('isFolder', file['gitObjectType'] == 'tree')]
                        file_sizes = get_file_size(server_url, project, repo_id, sha1_list, pat, api_version)

                        file_paths = [file['path'] for file in files if not file.get('isFolder', file['gitObjectType'] == 'tree')]
                        commit_counts = get_commit_count(server_url, project, repo_id, file_paths, pat, api_version)

                        for file in files:
                            is_folder = file.get('isFolder', file['gitObjectType'] == 'tree')
                            file_path = file['path']
                            sha1 = file['objectId'] if not is_folder else None
                            size = file_sizes.get(sha1, 0)
                            commit_count = commit_counts.get(file_path, 0)

                            file_info = {
                                'Collection Name': server_url.split('/')[-1],
                                'Project Name': project,
                                'Repository Name': repository_name,
                                'Branch Name': branch,
                                'File Name': file_path.split('/')[-1],
                                'File Type': 'Folder' if is_folder else 'File',
                                'Folder Level': file_path.count('/') - 1,
                                'Path': file_path,
                                'Size (Bytes)': int(size),
                                'Last Modified Time': last_modified,
                                'Author': author,
                                'Comments': comment,
                                'Commit ID': commit_id,
                                'Commit Count': commit_count
                            }
                            data_source_code.append(file_info)

                # Retrieve all commits in the repository
                all_commits = get_all_repo_commits(server_url, project, repo_id, pat, api_version, batch_size=batch_size)
                for commit in all_commits:
                    all_commit_info = {
                        'Author': commit['author']['name'],
                        'Commit Message': commit['comment'],
                        'Commit ID': commit['commitId'],
                        'Commit Date': commit['author']['date'],
                        'Tag Name': 'not tagged'  # Placeholder, will update if tags found
                    }
                    data_all_commits.append(all_commit_info)

                # Retrieve tags for the repository and fetch tag details in batch
                tags = get_tags(server_url, project, repo_id, pat, api_version, batch_size=batch_size)
                tag_ids = [tag['objectId'] for tag in tags]
                tag_details_map = get_tag_details(server_url, project, repo_id, tag_ids, pat, api_version)

                commit_tag_map = map_commit_tags(tags)  # Map for quick tag lookup by commit ID

                for tag in tags:
                    tag_id = tag['objectId']
                    tag_details = tag_details_map.get(tag_id)
                    if tag_details:
                        commit_details = get_commit_details(server_url, project, repo_id, tag_details['taggedObject']['objectId'], pat, api_version)
                        if commit_details:
                            tag_date, tag_time = tag_details['taggedBy']['date'].split('T')
                            tag_time = tag_time.split('Z')[0]
                            tag_info = {
                                'Tag Name': tag['name'].replace('refs/tags/', ''),
                                'Tag ID': tag['objectId'],
                                'Tag Message': tag_details['message'],
                                'Commit ID': tag_details['taggedObject']['objectId'],
                                'Commit Message': commit_details['comment'],
                                'Author': tag_details['taggedBy']['name'],
                                'Date & Time': f"{tag_date} {tag_time}"
                            }
                            data_tags.append(tag_info)

                # Update all_commit_info with tag names using commit_tag_map
                for commit in data_all_commits:
                    commit['Tag Name'] = commit_tag_map.get(commit['Commit ID'], 'not tagged')

                return data_source_code, data_commits, data_all_commits, data_tags

    # Return empty data if no repositories or data found
    return [], [], [], []



def construct_input(df):
    result = defaultdict(lambda: {"pat": "", "projects": []})

    for index, row in df.iterrows():
        if pd.isna(row['Server URL']) or pd.isna(row['PAT']):
            print(f"Skipping row {index + 1} due to missing data. ServerURL and PAT values are mandatory.")
            continue

        server_url = row['Server URL'].rstrip('/')
        pat = row['PAT']
        project_name = row.get('Project Name')
        repo_name = row.get('Repository Name')
        branch_name = row.get('Branch Name')

        # Update the PAT only if it's not already set for this server URL
        if not result[server_url]["pat"]:
            result[server_url]["pat"] = pat

        # Process row based on the presence of project name
        if project_name:
            # Check if the project already exists
            project = next((p for p in result[server_url]["projects"] if p["name"] == project_name), None)
            if project is None:
                project = {"name": project_name, "repos": []}
                result[server_url]["projects"].append(project)
        else:
            # If project name is empty, retrieve all projects and their repos
            for proj_name in get_project_names(devops_server_url=server_url, pat=pat):
                # Check if the project already exists
                project = next((p for p in result[server_url]["projects"] if p["name"] == proj_name), None)
                if not project:
                    project = {"name": proj_name, "repos": []}
                    result[server_url]["projects"].append(project)

                # Fetch repos for each project when project_name is empty
                for repo_name in get_repo_names_by_project(devops_server_url=server_url, pat=pat, project_name=proj_name):
                    repo = next((r for r in project["repos"] if r["name"] == repo_name), None)
                    if not repo:
                        repo = {"name": repo_name, "branches": []}
                        project["repos"].append(repo)
            continue  # Skip further processing for this row since we already handled all projects and repos

        # Prepare repository structure if project_name is specified
        if repo_name:
            # Check if repo already exists under this project
            repo = next((r for r in project["repos"] if r["name"] == repo_name), None)
            if not repo:
                # If repo does not exist, create it with branches
                repo = {"name": repo_name, "branches": []}
                project["repos"].append(repo)

            # Add branch if specified and unique
            if branch_name and branch_name not in repo["branches"]:
                repo["branches"].append(branch_name)
        else:
            # If repo is not specified, retrieve repos for the project
            for repo_name in get_repo_names_by_project(devops_server_url=server_url, pat=pat, project_name=project_name):
                repo = next((r for r in project["repos"] if r["name"] == repo_name), None)
                if not repo:
                    repo = {"name": repo_name, "branches": []}
                    project["repos"].append(repo)

    # Convert default dict to normal dict for final output
    return dict(result)


def main():
    input_file = r'git_discovery_input_form.xlsx'
    try:
        run_id = str(int(datetime.now().strftime("%Y%m%d%H%M%S")))
        output_directory = os.path.join("Git", run_id)

        # Create output directory if it doesn't exist
        if not os.path.exists(output_directory):
            os.makedirs(output_directory)
            print(f"Output directory created: {output_directory}")

        df = pd.read_excel(input_file)
        # Read the values from the Excel file and strip any leading/trailing spaces
        df['Server URL'] = df['Server URL'].fillna('').str.strip()
        df['Project Name'] = df['Project Name'].fillna('').str.strip()
        df['PAT'] = df['PAT'].fillna('').str.strip()
        df['Repository Name'] = df['Repository Name'].fillna('').str.strip()
        df['Branch Name'] = df['Branch Name'].fillna('').str.strip()

        # Form the input data in below format
        # Sample format below
        """
            {
               "server_url":{
                  "pat":"123test_token",
                  "projects":[
                     {
                        "name":"proj_name",
                        "repos":[
                           {
                              "name":"repo_name",
                              "branches":["branch_name_1", "branch_name_2"]
                           }
                        ]
                     }
                  ]
               }
            }
        """
        input_data = construct_input(df)
        print(f"Final input combination: {input_data}")

        for server_url, server_data in input_data.items():
            start_time = datetime.now()
            pat = server_data["pat"]
            for project in server_data["projects"]:
                proj_name = project["name"]
                print(f"Processing project {proj_name}")
                try:
                    for repo in project["repos"]:
                        master_data_source_code = []
                        master_data_commits = []
                        master_data_all_commits = []
                        master_data_tags = []
                        repo_name = repo["name"]
                        branches = repo.get("branches", [])
                        # If branches exist, iterate through them; otherwise, use an empty string for branch_name
                        for branch_name in branches if branches else [None]:
                            data_source_code, data_commits, data_all_commits, data_tags = process(server_url, pat, proj_name, repo_name, branch_name)
                            master_data_source_code = master_data_source_code + data_source_code
                            master_data_commits = master_data_commits + data_commits
                            master_data_all_commits = master_data_all_commits + data_all_commits
                            master_data_tags = master_data_tags + data_tags
                        file_id = str(int(datetime.now().strftime("%Y%m%d%H%M%S")))
                        output_filename = f"{proj_name}_{repo_name}__git_discovery_report_{file_id}.xlsx"
                        output_path = os.path.join(output_directory, output_filename)
                        generate_report(master_data_source_code, master_data_commits, master_data_all_commits, master_data_tags,output_path, proj_name, repo_name, server_url, start_time)
                        print(f'Report generated: {output_path}')

                        # Clear data and enforce garbage collection after each repository's report is generated
                        master_data_source_code.clear()
                        master_data_commits.clear()
                        master_data_all_commits.clear()
                        master_data_tags.clear()
                        gc.collect()  # Enforce garbage collection
                        gc.collect()
                except Exception as e:
                    logger.error(f"Error occurred while processing project '{proj_name}': {e}")
    except Exception as e:
        logger.error(f"Error occurred while processing input file '{input_file}': {e}")
        return


if __name__ == "__main__":
    main()