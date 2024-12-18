import os
from datetime import datetime
import openpyxl
import tempfile
import shutil
from urllib.parse import quote
import time
import subprocess
from requests.auth import HTTPBasicAuth
import pytz
import sys
import requests
from wiki_comments_discover import get_wiki_comments, get_page_id
from wiki_migrate_get_db import db_get_wiki
db_get_wiki()

# Add the 'src' directory to the system path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))
from src.wiki.wiki_db import db_post_wiki
from src.wiki.wiki_discover_comments_db import db_post_wiki as db_post_comments_wiki


def encode_url_component(component):
    return quote(component, safe='')


def get_last_modified_date_from_git(file_path, repo_dir):
    """
    Get the last commit date for a file in the Git repository.
    """
    try:
        # Run the git log command to get the last commit date for the file
        command = ["git", "log", "-1", "--format=%cd", file_path]
        result = subprocess.run(command, cwd=repo_dir, capture_output=True, text=True)
        
        if result.returncode == 0:
            # Get the commit date from the output (format: Mon Nov 23 14:38:56 2020 +0000)
            commit_date_str = result.stdout.strip()
            # Convert to datetime object
            last_modified = datetime.strptime(commit_date_str, "%a %b %d %H:%M:%S %Y %z")
            
            # Convert to IST (Indian Standard Time)
            ist = pytz.timezone('Asia/Kolkata')
            last_modified = last_modified.astimezone(ist)
            return last_modified.strftime('%Y-%m-%d %H:%M:%S')
        else:
            print(f"Error fetching last commit date for {file_path}: {result.stderr}")
            return None
    except Exception as e:
        print(f"Error executing git log for {file_path}: {e}")
        return None


def clone_wiki_repo(auth_clone_url):
    temp_dir = tempfile.mkdtemp()
    try:
        print(f"Cloning repository from {auth_clone_url} into {temp_dir}...")
        command = f"git clone {auth_clone_url} {temp_dir}"
        process = os.popen(command)
        output = process.read()
        process.close()
        if "fatal:" in output or "error:" in output:
            print(f"Error during git clone: {output}")
            return None
        else:
            print(f"Repository successfully cloned into {temp_dir}")
            return temp_dir
    except Exception as e:
        print(f"Error during git clone: {e}")
        return None


def discover_wiki_pages(wiki_repo_path):
    markdown_files = []
    for root, _, files in os.walk(wiki_repo_path):
        for file in files:
            if file.endswith(".md"):
                file_path = os.path.join(root, file)
                relative_path = os.path.relpath(file_path, wiki_repo_path)
                
                # Get the last modified date using Git
                last_modified = get_last_modified_date_from_git(relative_path, wiki_repo_path)
                
                if last_modified:
                    file_size = os.path.getsize(file_path)
                    markdown_files.append((relative_path, file_size, last_modified))
    
    return markdown_files


def handle_remove_readonly(func, path, excinfo):
    try:
        os.chmod(path, 0o777)  # Change the file mode to make it writable
        func(path)  # Retry the removal of the file
    except Exception as e:
        print(f"Error removing {path}: {e}")


def get_project_and_wiki_id(server_url, project_name, username, password):
    # Construct the full API URL
    api_url = f"{server_url}/{project_name}/_apis/wiki/wikis/{project_name}.wiki?api-version=6.0"
    
    # Send a GET request to the API with basic authentication
    response = requests.get(api_url, auth=HTTPBasicAuth(username, password))
    
    if response.status_code == 200:
        data = response.json()
        project_id = data.get('projectId')
        wiki_id = data.get('id')
        
        return project_id, wiki_id
    else:
        print(f"Error: Unable to fetch data (Status code: {response.status_code})")
        return None, None


def main():
    current_dir = os.path.dirname(os.path.abspath(__file__))

    input_file = os.path.join(current_dir, "wiki_migrate_input.xlsx")

    if not os.path.exists(input_file):
        raise FileNotFoundError(f"Input file not found at: {input_file}")
    
    username = input("Enter your username: ").strip()  
    password = input("Enter your Password: ").strip()  
    encoded_password = encode_url_component(password)

    # Load the input Excel file
    wb_input = openpyxl.load_workbook(input_file)
    ws_input = wb_input.active
    
    # Create a new workbook for the output report
    wb_report = openpyxl.Workbook()
    ws_report = wb_report.active
    ws_report.append(["Project Name", "File Path", "File Size (bytes)", "Last Modified", "Page ID", "Project ID", "Wiki ID"])

    # Create a separate sheet for comments
    ws_comments = wb_report.create_sheet("Comments")
    ws_comments.append(["Project Name", "File Path", "Comment ID", "Comment Text", "Created By", "Created Date"])

    for row in range(2, ws_input.max_row + 1):
        # Read data from the Excel file
        server_url = ws_input.cell(row=row, column=2).value.strip() if ws_input.cell(row=row, column=2).value else ''
        project_name = ws_input.cell(row=row, column=3).value.strip() if ws_input.cell(row=row, column=3).value else ''
        wiki_path = f"{project_name}.wiki"
        pat = ws_input.cell(row=row, column=4).value.strip() if ws_input.cell(row=row, column=2).value else ''
        
        if not server_url or not project_name or not wiki_path:
            print(f"Skipping row {row}: missing necessary data.")
            continue
        
        # Construct the URL for cloning
        clone_url = f"{server_url}/{project_name}/_git/{wiki_path}"
        auth_clone_url = clone_url.replace("://", f"://{username}:{encoded_password}@")
        
        # Cloning and processing the wiki
        temp_dir = clone_wiki_repo(auth_clone_url)
        if temp_dir:
            wiki_pages = discover_wiki_pages(temp_dir)
            if wiki_pages:
                for file in wiki_pages:
                    file_path, size, modified = file
                    project_id, wiki_id = get_project_and_wiki_id(server_url, project_name, username, pat)
                    
                    page_id = get_page_id(server_url, project_id, wiki_id, username, pat, f"/{file_path[:-3].replace('-', '%20')}")
                    # Create a data dictionary for the db_post_wiki function
                    data = {
                        "collection_name":server_url.split('/')[-1],
                        "project_id" : project_id,
                        "project_name": project_name,
                        "wiki_id": wiki_id,
                        "file_path": file_path[:-3],
                         "page_id": page_id,
                        "size_bytes": size,
                        "last_modified": modified
                    }
                    
                    comments = get_wiki_comments(server_url, project_id, wiki_id, page_id, username, pat)

                    # Store comment data in the Comments sheet
                    if comments:
                        for comment in comments:
                            comment_id = comment.get('comment_id', 'N/A')
                            comment_text = comment.get('comment_text', 'No Text')
                            created_by = comment.get('created_by', 'Unknown')
                            created_date = comment.get('created_date', 'Unknown Date')

                            # Append the comment data to the comments sheet
                            ws_comments.append([project_name, file_path[:-3], comment_id, comment_text, created_by, created_date])

                            # Create a data dictionary for the db_post_wiki_comments function
                            data_comments = {
                                "collection_name": server_url.split('/')[-1],
                                "project_name": project_name,
                                "file_path": file_path[:-3],
                                "comment_id": comment_id,
                                "comment_text": comment_text,
                                "created_by": created_by,
                                "created_date": created_date
                            }
                            
                            db_post_comments_wiki(data_comments)
                          
                            

                    else:
                        print(f"No comments found for Page ID: {page_id}")
                    
                    # Write data to the database
                    try:
                        db_post_wiki(data)
                       
                        print(f"Data successfully written to the database: {data}")
                    except Exception as e:
                        print(f"Error writing to database for file {file_path}: {e}")
                    
                    # Append data to the main Excel report sheet
                    ws_report.append([project_name, file_path[:-3], size, modified, page_id, project_id, wiki_id])
            else:
                print(f"No wiki pages found for project: {project_name}.")
            
            time.sleep(1)
            shutil.rmtree(temp_dir, onerror=handle_remove_readonly)
            print(f"Temporary directory {temp_dir} cleaned up.")
        else:
            print(f"Failed to clone the repository for project {project_name}. Skipping.")
    
    # Save the Excel report
    wb_report.save("wiki_source_discovery_reports.xlsx")
    print("Excel report generated: wiki_source_discovery_reports.xlsx")
    
    print("All data has been processed and written to the database.")


if __name__ == "__main__":
    main()
