import pandas as pd
import os
import sys
from sqlalchemy.orm import sessionmaker
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import PatternFill
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))
from src.dbDetails.db import get_db
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def create_reconciliation_folder():
    folder_path = os.path.join(os.getcwd(), "Reconciliation")  
    if not os.path.exists(folder_path):  
        os.makedirs(folder_path)  
    return folder_path

# Function to extract name from the URL
def extract_name(url):
    return url.split("/")[-1]  

# Function to fetch work item counts from source and target tables
def fetch_work_item_counts(db_session, project_name):
    # Query source table for work item counts
    source_query = text(f"""
        SELECT workitem_type, COUNT(*) 
        FROM devops_to_ados.db_devops_discovery_boards_workitem_details
        WHERE project_name = :project_name
        GROUP BY workitem_type
    """)
    source_counts = db_session.execute(source_query, {"project_name": project_name}).fetchall()

    # Query target table for work item counts
    target_query = text(f"""
        SELECT workitem_type, COUNT(*) 
        FROM devops_to_ados.db_ado_discovery_boards_workitem_details
        WHERE project_name = :project_name
        GROUP BY workitem_type
    """)
    target_counts = db_session.execute(target_query, {"project_name": project_name}).fetchall()

    # Convert counts into dictionaries for easy lookup
    source_dict = {row[0]: row[1] for row in source_counts}
    target_dict = {row[0]: row[1] for row in target_counts}

    return source_dict, target_dict

# Function to create Excel file for each project
def generate_excel_for_project(db_session, source_server_url, target_organization_url, source_project_name, target_project_name):
    # Extract names from URLs
    source_collection_name = extract_name(source_server_url)
    target_organization_name = extract_name(target_organization_url)

    # Fetch work item counts
    source_counts, target_counts = fetch_work_item_counts(db_session, source_project_name)

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f'{source_project_name}_Comparison'

    # Create header row
    header = [
        "Source Collection", "Source Project", "Source Work Item Type", "Source Work Item Count",
        "Target Organization", "Target Project", "Target Work Item Type", "Target Work Item Count", "Status"
    ]
    ws.append(header)

    # Define fill colors for status
    header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")  # Navy blue
    matched_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
    not_matched_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Pale red

    # Apply header styles
    for col in range(1, len(header) + 1):
        header_cell = ws.cell(row=1, column=col)
        header_cell.fill = header_fill
        header_cell.font = Font(color="FFFFFF", bold=True)  # White text with bold font

    # Merge source and target work item types
    all_work_item_types = set(source_counts.keys()).union(set(target_counts.keys()))

    # Compare work item types and write data
    for work_item_type in all_work_item_types:
        source_count = source_counts.get(work_item_type, 0)
        target_count = target_counts.get(work_item_type, 0)
        status = "MATCHED" if source_count == target_count else "NOT MATCHED"
        row = [
            source_collection_name, source_project_name, work_item_type, source_count,
            target_organization_name, target_project_name, work_item_type, target_count, status
        ]
        ws.append(row)

        # Apply status color
        status_cell = ws.cell(row=ws.max_row, column=9)  # Status is in the 9th column
        if status == "MATCHED":
            status_cell.fill = matched_fill
        else:
            status_cell.fill = not_matched_fill

    # Adjust column widths to ensure the data fits properly and is evenly spaced
    for col in range(1, len(header) + 1):
        max_length = 0
        column = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    # Save the workbook as an Excel file
    folder_path = create_reconciliation_folder()
    file_name = f'{source_project_name}_Reconciliation_report.xlsx'
    file_path = os.path.join(folder_path, file_name)
    wb.save(file_path)
    print(f"Excel file generated for project: {source_project_name} at {file_name}")

# Main function to generate Excel files for all projects in the db_devops_ado_migration_details table
def generate_excel_files_for_all_projects():
    db_session = next(get_db())
    try:   
        migration_query = text("""
            SELECT source_server_url, source_project_name, target_organization_url, target_project_name
            FROM devops_to_ados.db_devops_ado_migration_details
        """)
        migration_rows = db_session.execute(migration_query).fetchall()

        # Generate Excel for each project
        for row in migration_rows:
            source_server_url, source_project_name, target_organization_url, target_project_name = row
            generate_excel_for_project(db_session, source_server_url, target_organization_url, source_project_name, target_project_name)
    finally:
     db_session.close()

if __name__ == "__main__":
    generate_excel_files_for_all_projects()
