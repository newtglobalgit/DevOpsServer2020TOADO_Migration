import os
import re
import shutil
import subprocess
import time
import pandas as pd
import requests
from requests.auth import HTTPBasicAuth
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import getpass
import logging
import git
from git import Repo
from urllib.parse import quote
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from collections import defaultdict
from utils.common import get_project_names, get_repo_names_by_project

log_dir = "logs"
if not os.path.exists(log_dir):
    os.makedirs(log_dir)
# Create a logger
logger = logging.getLogger()
# Set the log level
logger.setLevel(logging.INFO)
# File handler
file_handler = logging.FileHandler(os.path.join(log_dir, f'git_discovery_{datetime.now().strftime("%Y%m%d%H%M%S")}.log'))
file_handler.setLevel(logging.INFO)
# Console handler
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
# Log format
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)
# Add handlers to the logger
logger.addHandler(file_handler)
logger.addHandler(console_handler)


def make_request_with_retries(url, auth, max_retries=10, timeout=300):
    logger.info(f"Making request to URL: {url}")
    for attempt in range(max_retries):
        try:
            response = requests.get(url, auth=auth, timeout=timeout)
            if response.status_code == 200:
                logger.info(f"Request successful with status code: {response.status_code}")
                return response
            else:
                logger.warning(f"Issue during request: {response.text}")
                logger.warning(f"Attempt {attempt + 1}: Request timed out. Retrying...")
                time.sleep(2 ** attempt)
        except Exception as e:
            logger.error(f"Error during request: {e}")
            break
    logger.error(f"Failed to connect after {max_retries} attempts.")
    return None

def modify_value(value):
    try:
        # Replacement dictionary for special characters
        replacements = {
            '+': '%2B', '%24': '$', '&': '%26', ',': '%2C', ':': '%3A',
            ';': '%3B', '=': '%3D', '?': '%3F', '@': '%40', ' ': '%20',
            '"': '%22', '<': '%3C', '>': '%3E', '#': '%23', '{': '%7B',
            '}': '%7D', '|': '%7C', '\\': '%5C', '^': '%5E', '[': '%5B',
            ']': '%5D', '`': '%60'
        }

        # Log the original password (for debugging purposes, make sure this is secure in production!)
        logger.info(f"Original password: XXXXXX")

        # Replace all special characters in the password
        for char, replacement in replacements.items():
            value = value.replace(char, replacement)
        
        # Log the modified password
        logger.info(f"Modified password: XXXXX")
        return value
    
    except Exception as e:
        logger.error(f"Failed to modify password: {e}")
        return value

def encode_url_component(component):
    return quote(component, safe='')


def make_request_with_retries(url, pat, method='GET', data=None, timeout=300, max_retries=10, backoff_factor=0.5):
    auth = HTTPBasicAuth('', pat)
    session = requests.Session()
    retries = Retry(
        total=max_retries,
        backoff_factor=backoff_factor,
        status_forcelist=[408, 429, 500, 502, 503, 504],
        allowed_methods=["HEAD", "GET", "OPTIONS", "POST"]
    )
    session.mount('https://', HTTPAdapter(max_retries=retries))
    for attempt in range(max_retries):
        try:
            if method == 'GET':
                response = session.get(url, auth=auth, timeout=timeout)
            elif method == 'POST':
                response = session.post(url, auth=auth, json=data, timeout=timeout)
            else:
                logger.error(f"HTTP method '{method}' is not supported.")
                return None
            if response.status_code == 200:
                logger.info(f"Request succeeded for URL {url}")
                return response
            elif response.status_code == 409:
                logger.warning("Received 409 Conflict. Retrying after delay...")
            elif response.status_code == 404:
                logger.info(f"Resource not found (404) for URL {url}.")
                return None
            elif response.status_code == 403:
                logger.error("Access forbidden (403). Check permissions or credentials.")
                return None
            else:
                logger.warning(f"Unexpected status code {response.status_code} for URL {url}.")
                return None
        except requests.exceptions.RequestException as e:
            logger.error(f"Error occurred while making request to {url}: {e}")
            return None
    logger.error(f"Exceeded max retries for URL {url}")
    return None


# Function to read configuration from Excel file
def read_config_from_excel(file_path):
    try:
        df = pd.read_excel(file_path, sheet_name='Input')
        print("DataFrame loaded successfully.")
        print("DataFrame content:")
        print(df)
    except ValueError as e:
        print(f"Error: {e}")
        return None, None, None, None, None, None

    server_url = df.at[0, 'Server URL'].strip()
    project = df.at[0, 'Project Name'].strip()
    pat = df.at[0, 'PAT'].strip()
    repository_name = df.at[0, 'Repository Name'].strip()
    branch_name = df.at[0, 'Branch Name'].strip() if not pd.isna(df.at[0, 'Branch Name']) else None

    if not server_url or not project or not pat or not repository_name:
        print("Please ensure the Excel file contains the required data.")
        return None, None, None, None, None, None

    return server_url, project, pat, repository_name, branch_name, df

    
def authenticate_and_get_projects(server_url, pat, api_version):
    url = f'{server_url}/_apis/projects?api-version={api_version}'
    response = make_request_with_retries(url, pat)
    if response:
        logger.info('Login successful.')
        return response.json()
    else:
        logger.error('Failed to authenticate.')
        return None


def get_repositories(server_url, project, pat, api_version):
    # Encode project name for safe URL usage
    encoded_project = requests.utils.quote(project)
    url = f'{server_url}/{encoded_project}/_apis/git/repositories?api-version={api_version}'

    # Make the request
    response = requests.get(url, auth=HTTPBasicAuth('', pat))

    if response.status_code == 200:
        # Extract repositories from response
        all_repositories = response.json().get('value', [])
        return {"value": all_repositories}
    else:
        # Log error and return empty result
        logger.error(f"Failed to retrieve repositories: {response.status_code}, {response.text}")
        return {"value": []}


def get_branches(server_url, project, repository_id, pat, api_version):
    # Encode project and repository ID for safe URL usage
    encoded_project = requests.utils.quote(project)
    encoded_repository_id = requests.utils.quote(repository_id)

    # Construct the URL without batch size
    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/refs?filter=heads&api-version={api_version}'

    # Make the request
    response = requests.get(url, auth=HTTPBasicAuth('', pat))

    if response.status_code == 200:
        # Extract branches from response
        all_branches = response.json().get('value', [])
        return all_branches
    else:
        # Log error and return an empty list
        logger.error(f"Failed to retrieve branches: {response.status_code}, {response.text}")
        return []


def get_latest_commit_info(server_url, project, repository_id, branch_name, pat, api_version):
    encoded_project = encode_url_component(project)
    encoded_repository_id = encode_url_component(repository_id)
    encoded_branch_name = encode_url_component(branch_name)
    
    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/commits?searchCriteria.itemVersion.version={encoded_branch_name}&$top=1&api-version={api_version}'
    response = make_request_with_retries(url, pat, method='GET')

    if response:
        commit = response.json().get('value', [None])[0]
        if commit:
            return commit['commitId'], commit['comment'], commit['author']['name'], commit['author']['date']
    logger.error('Failed to retrieve latest commit info.')
    return None, None, None, None


def get_files_in_branch(server_url, project, repository_id, branch_name, pat, api_version, batch_size=150):
    encoded_project = encode_url_component(project)
    encoded_repository_id = encode_url_component(repository_id)
    encoded_branch_name = encode_url_component(branch_name)
    
    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/items?scopePath=/&recursionLevel=Full&versionDescriptor[version]={encoded_branch_name}&$top={batch_size}&api-version={api_version}'
    all_files = []
    continuation_token = None
    last_token = None  # To detect repetitive tokens
    repeat_count = 0   # Count for repeated tokens to avoid infinite loops
    
    while True:
        headers = {'x-ms-continuationtoken': continuation_token} if continuation_token else {}
        response = make_request_with_retries(url, pat, method='GET')

        if response:
            data = response.json().get('value', [])
            all_files.extend(data)
            
            # Update continuation token
            last_token = continuation_token
            continuation_token = response.headers.get('x-ms-continuationtoken')
            
            # Logging for continuation token
            logger.info(f"Continuation token for next batch: {continuation_token}")
            
            # Check if the continuation token is the same as the last one
            if continuation_token == last_token:
                repeat_count += 1
                if repeat_count > 3:  # Break after 3 repeated tokens
                    logger.warning(f"Breaking due to repeated continuation tokens for branch {branch_name}.")
                    break
            else:
                repeat_count = 0  # Reset repeat count if the token changes
            
            # Exit loop if there is no continuation token
            if not continuation_token:
                logger.info(f"Finished retrieving all items for branch {branch_name}.")
                break
        else:
            logger.error('Failed to retrieve items in branch with pagination.')
            break

    return all_files


def get_file_size(server_url, project, repository_id, sha1, pat, api_version):
    encoded_project = encode_url_component(project)
    encoded_repository_id = encode_url_component(repository_id)
    encoded_sha1 = encode_url_component(sha1)
    
    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/blobs/{encoded_sha1}?api-version={api_version}'
    response = make_request_with_retries(url, pat, method='GET')

    if response:
        return response.headers.get('Content-Length', 0)
    logger.error('Failed to retrieve blob size.')
    return 0


def get_commit_count(server_url, project, repository_id, file_path, pat, api_version):
    encoded_project = encode_url_component(project)
    encoded_repository_id = encode_url_component(repository_id)
    encoded_file_path = encode_url_component(file_path)
    
    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/commits?searchCriteria.itemPath={encoded_file_path}&api-version={api_version}'
    response = make_request_with_retries(url, pat, method='GET')

    if response:
        return max(response.json().get('count', 1), 1)
    logger.error(f'Failed to retrieve commit count for {file_path}.')
    return 1

    
def get_all_commits(server_url, project, repository_id, branch_name, pat, api_version):
    # Encode project, repository ID, and branch name for safe URL usage
    encoded_project = requests.utils.quote(project)
    encoded_repository_id = requests.utils.quote(repository_id)
    encoded_branch_name = requests.utils.quote(branch_name)

    # Construct the URL without batch size
    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/commits?searchCriteria.itemVersion.version={encoded_branch_name}&api-version={api_version}'

    # Make the request
    response = requests.get(url, auth=HTTPBasicAuth('', pat))

    if response.status_code == 200:
        # Extract commits from response
        all_commits = response.json().get('value', [])
        return all_commits
    else:
        # Log error and return an empty list
        logger.error(f"Failed to retrieve commits: {response.status_code}, {response.text}")
        return []


def clone_repo_branch(repo_url, destination_path, pat, branch_name, username, password):
    try:
        # Prepare the URL with authentication (PAT)
        if repo_url.startswith('http://'):
            auth_repo_url = repo_url.replace('http://', f'http://{username}:{password}@')
        else:
            auth_repo_url = repo_url.replace('https://', f'https://{username}:{password}@')

        logger.info(f'Cloning started {datetime.now()}')

        # Use subprocess to clone the repository with the specified branch
        command = f'git clone --branch {branch_name} --single-branch {auth_repo_url} {destination_path}'

        print(command)
        # Run the clone command
        print(f"Cloning branch '{branch_name}' from repository at {repo_url}...")
        result = subprocess.run(command, shell=True, capture_output=True, text=True)

        # Check for errors in the cloning process
        if result.returncode != 0:
            print(f"Error cloning repository: {result.stderr}")
            logger.error(f"Error cloning repository: {result.stderr}")
            return

        print(f"Branch '{branch_name}' successfully cloned to {destination_path}")
        logger.info(f'Cloning completed {datetime.now()}')

    except Exception as e:
        print(f"An error occurred while cloning the repository branch: {e}")
        logger.error(f"An error occurred while cloning the repository branch: {e}")

    
def get_all_repo_commits(server_url, project, repository_id, pat, api_version, batch_size=150):
    # Encode project and repository ID for safe URL usage
    encoded_project = requests.utils.quote(project)
    encoded_repository_id = requests.utils.quote(repository_id)

    # Construct the URL
    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/commits?$top={batch_size}&api-version={api_version}'

    # Make the request
    response = requests.get(url, auth=HTTPBasicAuth('', pat))

    if response.status_code == 200:
        # Extract commits from response
        all_commits = response.json().get('value', [])
        return all_commits
    else:
        # Log error and return an empty list
        logger.error(f"Failed to retrieve commits: {response.status_code}, {response.text}")
        return []


def get_tags(server_url, project, repository_id, pat, api_version):
    # Encode project and repository ID for safe URL usage
    encoded_project = requests.utils.quote(project)
    encoded_repository_id = requests.utils.quote(repository_id)

    # Construct the URL without batch size
    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/refs?filter=tags&api-version={api_version}'

    # Make the request
    response = requests.get(url, auth=HTTPBasicAuth('', pat))

    if response.status_code == 200:
        # Extract tags from response
        all_tags = response.json().get('value', [])
        return all_tags
    else:
        # Log error and return an empty list
        logger.error(f"Failed to retrieve tags: {response.status_code}, {response.text}")
        return []


def get_tag_details(server_url, project, repository_id, tag_id, pat, api_version):
    encoded_project = encode_url_component(project)
    encoded_repository_id = encode_url_component(repository_id)
    encoded_tag_id = encode_url_component(tag_id)
    
    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/annotatedtags/{encoded_tag_id}?api-version=6.0-preview.1'
    response = make_request_with_retries(url, pat, method='GET')

    if response:
        return response.json()
    logger.error(f'Failed to retrieve tag details for {tag_id}.')
    return None

def get_commit_details(server_url, project, repository_id, commit_id, pat, api_version):
    encoded_project = encode_url_component(project)
    encoded_repository_id = encode_url_component(repository_id)
    encoded_commit_id = encode_url_component(commit_id)
    
    url = f'{server_url}/{encoded_project}/_apis/git/repositories/{encoded_repository_id}/commits/{encoded_commit_id}?api-version={api_version}&$top=100000'
    response = make_request_with_retries(url, pat, method='GET')

    if response:
        return response.json()
    logger.error(f'Failed to retrieve commit details for {commit_id}.')
    return None


def map_commit_tags(tags):
    commit_tag_map = {}
    for tag in tags:
        commit_id = tag['peeledObjectId'] if 'peeledObjectId' in tag else tag['objectId']
        commit_tag_map[commit_id] = tag['name']
    return commit_tag_map


def apply_header_styles(workbook, sheet_name):
    sheet = workbook[sheet_name]
    header_fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF') 
    for cell in sheet[1]:  
        cell.fill = header_fill   
        cell.font = header_font


def apply_black_border(sheet):
    thin_border = Border(left=Side(style='thin', color='000000'),
                         right=Side(style='thin', color='000000'),
                         top=Side(style='thin', color='000000'),
                         bottom=Side(style='thin', color='000000'))
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border


def remove_gridlines(sheet):
    sheet.sheet_view.showGridLines = False


def adjust_column_width(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Get the column name
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
                cell.alignment = Alignment(wrap_text=True)
            except:
                pass
        adjusted_width = min(max_length, 30)
        sheet.column_dimensions[column].width = adjusted_width


def align_cells(sheet):
    for row in sheet.iter_rows(min_row=2):  # Skip header row
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')
            elif isinstance(cell.value, datetime) or isinstance(cell.value, str) and '-' in cell.value and ':' in cell.value:
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')


def sanitize_sheet_name(name):
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    for char in invalid_chars:
        name = name.replace(char, '')
    return name[:31]  # Excel sheet names can have a maximum of 31 characters


def generate_report(data_source_code, data_commits, data_all_commits, data_tags, output_path, project_name, repo_name,
                    server_url, start_time):
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Create an empty summary sheet
        pd.DataFrame().to_excel(writer, sheet_name='summary', index=False)
        df_source_code = pd.DataFrame(data_source_code)
        df_commits = pd.DataFrame(data_commits)
        df_all_commits = pd.DataFrame(data_all_commits)
        df_tags = pd.DataFrame(data_tags)
        df_source_code.to_excel(writer, sheet_name='source_code', index=False)
        df_commits.to_excel(writer, sheet_name='commits', index=False)
        df_tags.to_excel(writer, sheet_name='tags', index=False)
    
    workbook = load_workbook(output_path)
    worksheet_summary = workbook['summary']
    apply_header_styles(workbook, 'source_code')
    apply_header_styles(workbook, 'commits')
    apply_header_styles(workbook, 'tags')

    apply_black_border(workbook['summary'])
    apply_black_border(workbook['source_code'])
    apply_black_border(workbook['commits'])
    apply_black_border(workbook['tags'])

    remove_gridlines(workbook['summary'])
    remove_gridlines(workbook['source_code'])
    remove_gridlines(workbook['commits'])
    remove_gridlines(workbook['tags'])



        # Make header text bold for all sheets except the summary sheet
    for sheet_name in ['source_code', 'commits', 'tags']:
        sheet = workbook[sheet_name]
        apply_header_styles(workbook, sheet_name)
        apply_black_border(sheet)
        remove_gridlines(sheet)
        adjust_column_width(sheet)
        align_cells(sheet)

    sheet_source_code = workbook['source_code']
    for row in sheet_source_code.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True) 

    run_duration = datetime.now() - start_time
    hours, remainder = divmod(run_duration.total_seconds(), 3600)
    minutes, seconds = divmod(remainder, 60)
    formatted_run_duration = f"{int(hours)} hours, {int(minutes)} minutes, {seconds:.1f} seconds"

    # Determine organization or collection based on the server URL
    if "dev.azure.com" in server_url:
        org_or_collection = "organization"
        org_or_collection_name = server_url.split("/")[-1]  # Extract organization name from URL
    else:
        org_or_collection = "collection"
        org_or_collection_name = server_url.split('/')[-1]  # Extract collection name from URL

    # Write summary data
    summary_data = {
        'Report Title': f"Project {project_name} Git Report.",
        'Purpose of the report': f"This provides a detailed view of the repo {repo_name} in project"
                                 f" {project_name} of {org_or_collection} {org_or_collection_name}.",
        'Run Date': datetime.now().strftime('%d-%b-%Y %I:%M %p'),
        'Run Duration': formatted_run_duration,
        'Run By': getpass.getuser(),
        'Input': f"Server URL: {server_url}, Project Name: {project_name}, Repository Name: {repo_name}"
    }

    row = 1
    for key, value in summary_data.items():
        worksheet_summary.cell(row=row, column=1).value = key
        worksheet_summary.cell(row=row, column=1).fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')
        worksheet_summary.cell(row=row, column=1).font = Font(color='FFFFFF', bold=True)
        worksheet_summary.cell(row=row, column=2).value = value
        worksheet_summary.cell(row=row, column=2).alignment = Alignment(horizontal='left')
        row += 1

    for i in range(len(summary_data)):
        worksheet_summary.column_dimensions['A'].width = 26.71  # Adjust column A width
        worksheet_summary.column_dimensions['B'].width = 85.87  # Adjust column B width
        worksheet_summary.row_dimensions[i + 1].height = 30  # Adjust row height

    apply_black_border(worksheet_summary)  # Apply black border to summary sheet

    workbook.save(output_path)

def remove_readonly(func, path, exc_info):
    """Clear the read-only flag before attempting to delete."""
    os.chmod(path, 0o777)  # Change the permissions to make it writable
    func(path)

def clean_local_clone_path(local_clone_path):
    try:
        # Check if the directory exists
        if os.path.exists(local_clone_path):
            logger.info(f"Force cleaning up the existing clone folder at {local_clone_path}")

            # Force delete by changing permissions and using shutil.rmtree
            shutil.rmtree(local_clone_path, onerror=remove_readonly)
            logger.info("Clone folder cleaned up successfully.")
        else:
            # If the directory doesn't exist, create it
            os.makedirs(local_clone_path)
            logger.info(f"Clone folder {local_clone_path} created successfully.")
    except PermissionError:
        logger.error(f"Permission denied while trying to clean or create the folder: {local_clone_path}")
    except Exception as e:
        logger.error(f"An error occurred while cleaning the folder: {e}")


def process(server_url, pat, project, repository_name, branch_name, username, password):

    try:
        api_version = '6.0'  # Change this if your server uses a different version


        # Print the server URL and project for debugging
        print(f"Server URL: {server_url}")
        print(f"Project: {project}")
        print(f"Repo: {repository_name}")
        print(f"Branch: {branch_name}")

        # Authenticate and get projects (optional step to confirm login)
        projects = authenticate_and_get_projects(server_url, pat, api_version)
        if projects:
            # Retrieve repository info
            repositories = get_repositories(server_url, project, pat, api_version)
            if repositories:
                for repo in repositories['value']:
                    logger.info(f"Repository Name : {repo.get('name')}")
                    if repo['name'] == repository_name:
                        repo_id = repo['id']

                        data_source_code = []
                        data_commits = []
                        data_all_commits = []
                        data_tags = []
                        repo_url = f"{server_url}/{project}/_git/{repo['name']}"
                        destination_path = os.path.join(cwd, 'clone', repo['name'])
                        if not os.path.exists(destination_path):
                            os.makedirs(destination_path)
                        # Get branches
                        
                        branches = get_branches(server_url, project, repo_id, pat, api_version)
                        if branches:
                            branch_names = [branch['name'].replace('refs/heads/', '') for branch in
                                            branches] if branch_name is None else [branch_name]

                            for branch in branch_names:
                                destination_path = os.path.join(cwd, 'clone', repo['name'], branch)
                                if not os.path.exists(destination_path):
                                    os.makedirs(destination_path)
                                clean_local_clone_path(os.path.join(cwd, 'clone'))
                                clone_repo_branch(repo_url, destination_path, pat, branch, username, password)
                                commit_id, comment, author, last_modified = get_latest_commit_info(server_url, project,
                                                                                                repo_id, branch, pat,
                                                                                                api_version)
                                if commit_id:
                                    # Get all commits for the branch
                                    commits = get_all_commits(server_url, project, repo_id, branch, pat, api_version)
                                    for commit in commits:
                                        commit_info = {
                                            'Collection Name': server_url.split('/')[-1],
                                            'Project Name': project,
                                            'Repository Name': repository_name,
                                            'Branch Name': branch,
                                            'Commit ID': commit['commitId'],
                                            'Commit Message': commit['comment'],
                                            'Author': commit['author']['name'],
                                            'Commit Date': commit['author']['date']
                                        }
                                        data_commits.append(commit_info)

                                    for dirpath, dirnames, filenames in os.walk(destination_path):
                                        if '.git' in dirpath:
                                            continue
                                        root_folder_name = os.path.basename(destination_path)
                                        project_folder_name = os.path.relpath(dirpath, destination_path)
                                        is_branch = "Yes" if dirpath in branches else "No"

                                        # Add folder details
                                        for dirname in dirnames:
                                            if '.git' in dirname:
                                                continue
                                            folder_path = os.path.join(dirpath, dirname)
                                            file_info = {
                                                    'Collection Name': server_url.split('/')[-1],
                                                    'Project Name': project,
                                                    'Repository Name': repository_name,
                                                    'Branch Name': branch,
                                                    'Name': dirname,
                                                    'File Type': 'Folder',
                                                    'Folder Level': project_folder_name.count('/') + 1,
                                                    'Path': folder_path,  
                                                    'Size (Bytes)': ""
                                                }
                                            data_source_code.append(file_info)

                                        # Add file details
                                        for filename in filenames:
                                            if '.git' in dirpath:
                                                continue
                                            file_path = os.path.join(dirpath, filename)
                                            if os.path.exists(file_path):
                                                file_size = os.path.getsize(file_path)
                                            else:
                                                file_size = 0
                                            file_type = os.path.splitext(filename)[1]

                                            file_info = {
                                                    'Collection Name': server_url.split('/')[-1],
                                                    'Project Name': project,
                                                    'Repository Name': repository_name,
                                                    'Branch Name': branch,
                                                    'Name': filename,
                                                    'File Type': 'File',
                                                    'Folder Level': file_path.count('/') + 1,
                                                    'Path': file_path,  
                                                    'Size (Bytes)': file_size
                                                }
                                            data_source_code.append(file_info)
                        # all_commits = get_all_repo_commits(server_url, project, repo_id, pat, api_version)
                        tags = get_tags(server_url, project, repo_id, pat, api_version)

                        for tag in tags:
                            tag_id = tag['objectId']
                            tag_details = get_tag_details(server_url, project, repo_id, tag_id, pat, api_version)
                            if tag_details:
                                commit_details = get_commit_details(server_url, project, repo_id,
                                                                    tag_details['taggedObject']['objectId'], pat,
                                                                    api_version)
                                if commit_details:
                                    tag_date, tag_time = tag_details['taggedBy']['date'].split('T')
                                    tag_time = tag_time.split('Z')[0]
                                    tag_info = {
                                        'Tag Name': tag['name'].replace('refs/tags/', ''),
                                        'Tag ID': tag['objectId'],
                                        'Tag Message': tag_details['message'],
                                        'Commit ID': tag_details['taggedObject']['objectId'],
                                        'Commit Message': commit_details['comment'],
                                        'Author': tag_details['taggedBy']['name'],
                                        'Date & Time': f"{tag_date} {tag_time}"
                                    }
                                    data_tags.append(tag_info)

                        # commit_tag_map = map_commit_tags(tags)
                        # for commit in all_commits:
                        #     tag_name = commit_tag_map.get(commit['commitId'], 'not tagged')
                        #     all_commit_info = {
                        #         'Author': commit['author']['name'],
                        #         'Commit Message': commit['comment'],
                        #         'Commit ID': commit['commitId'],
                        #         'Commit Date': commit['author']['date'],
                        #         'Tag Name': tag_name  # Include the Tag Name column
                        #     }
                        #     data_all_commits.append(all_commit_info)

                        return data_source_code, data_commits, data_all_commits, data_tags
        return [], [], [], []
    except Exception as e:
        logger.error(f"Error occurred while processing fetching details {e}")
        return
    


def construct_input(df):
    result = defaultdict(lambda: {"pat": "", "projects": []})

    for index, row in df.iterrows():
        if pd.isna(row['Server URL']) or pd.isna(row['PAT']):
            print(f"Skipping row {index + 1} due to missing data. ServerURL and PAT values are mandatory.")
            continue

        server_url = row['Server URL'].rstrip('/')
        pat = row['PAT']
        project_name = row.get('Project Name')
        repo_name = row.get('Repository Name')
        branch_name = row.get('Branch Name')

        # Update the PAT only if it's not already set for this server URL
        if not result[server_url]["pat"]:
            result[server_url]["pat"] = pat

        # Process row based on the presence of project name
        if project_name:
            # Check if the project already exists
            project = next((p for p in result[server_url]["projects"] if p["name"] == project_name), None)
            if not project:
                project = {"name": project_name, "repos": []}
                result[server_url]["projects"].append(project)
        else:
            # If project name is empty, retrieve all projects and their repos
            for proj_name in get_project_names(devops_server_url=server_url, pat=pat):
                # Check if the project already exists
                project = next((p for p in result[server_url]["projects"] if p["name"] == proj_name), None)
                if not project:
                    project = {"name": proj_name, "repos": []}
                    result[server_url]["projects"].append(project)

                # Fetch repos for each project when project_name is empty
                for repo_name in get_repo_names_by_project(devops_server_url=server_url, pat=pat, project_name=proj_name):
                    repo = next((r for r in project["repos"] if r["name"] == repo_name), None)
                    if not repo:
                        repo = {"name": repo_name, "branches": []}
                        project["repos"].append(repo)
            continue  # Skip further processing for this row since we already handled all projects and repos

        # Prepare repository structure if project_name is specified
        if repo_name:
            # Check if repo already exists under this project
            repo = next((r for r in project["repos"] if r["name"] == repo_name), None)
            if not repo:
                # If repo does not exist, create it with branches
                repo = {"name": repo_name, "branches": []}
                project["repos"].append(repo)

            # Add branch if specified and unique
            if branch_name and branch_name not in repo["branches"]:
                repo["branches"].append(branch_name)
        else:
            # If repo is not specified, retrieve repos for the project
            for repo_name in get_repo_names_by_project(devops_server_url=server_url, pat=pat, project_name=project_name):
                repo = next((r for r in project["repos"] if r["name"] == repo_name), None)
                if not repo:
                    repo = {"name": repo_name, "branches": []}
                    project["repos"].append(repo)

    # Convert default dict to normal dict for final output
    return dict(result)


def main(username, password):
    input_file = r'git_discovery_input_form.xlsx'
    username = modify_value(username)
    password = modify_value(password)
    try:
        run_id = str(int(datetime.now().strftime("%Y%m%d%H%M%S")))
        output_directory = os.path.join("Git", run_id)

        # Create output directory if it doesn't exist
        if not os.path.exists(output_directory):
            os.makedirs(output_directory)
            print(f"Output directory created: {output_directory}")

        df = pd.read_excel(input_file)

        # Read the values from the Excel file and strip any leading/trailing spaces
        df['Server URL'] = df['Server URL'].fillna('').str.strip()
        df['Project Name'] = df['Project Name'].fillna('').str.strip()
        df['PAT'] = df['PAT'].fillna('').str.strip()
        df['Repository Name'] = df['Repository Name'].fillna('').str.strip()
        df['Branch Name'] = df['Branch Name'].fillna('').str.strip()

        # Form the input data in below format
        # Sample format below
        """
            {
               "server_url":{
                  "pat":"123test_token",
                  "projects":[
                     {
                        "name":"proj_name",
                        "repos":[
                           {
                              "name":"repo_name",
                              "branches":["branch_name_1", "branch_name_2"]
                           }
                        ]
                     }
                  ]
               }
            }
        """
        input_data = construct_input(df)
        print(f"Final input combination: {input_data}")

        for server_url, server_data in input_data.items():
            pat = server_data["pat"]
            for project in server_data["projects"]:
                proj_name = project["name"]
                print(f"Processing project {proj_name}")
                try:
                    for repo in project["repos"]:
                        start_time = datetime.now()
                        master_data_source_code = []
                        master_data_commits = []
                        master_data_all_commits = []
                        master_data_tags = []
                        repo_name = repo["name"]
                        branches = repo.get("branches", [])
                        file_id = str(int(datetime.now().strftime("%Y%m%d%H%M%S")))
                        output_filename = f"{proj_name}_{repo_name}__git_discovery_report_{file_id}.xlsx"
                        output_path = os.path.join(output_directory, output_filename)
                        # If branches exist, iterate through them; otherwise, use an empty string for branch_name
                        for branch_name in branches if branches else [None]:
                            data_source_code, data_commits, data_all_commits, data_tags = process(server_url, pat, proj_name, repo_name, branch_name, username, password)
                            master_data_source_code = master_data_source_code + data_source_code
                            master_data_commits = master_data_commits + data_commits
                            master_data_all_commits = master_data_all_commits + data_all_commits
                            master_data_tags = master_data_tags + data_tags
                        file_id = str(int(datetime.now().strftime("%Y%m%d%H%M%S")))
                        output_filename = f"{proj_name}_{repo_name}__git_discovery_report_{file_id}.xlsx"
                        output_path = os.path.join(output_directory, output_filename)
                        generate_report(master_data_source_code, master_data_commits, master_data_all_commits, master_data_tags,
                                        output_path, proj_name, repo_name, server_url, start_time)
                        logger.info(f'Report generated: {output_path}')
                except Exception as e:
                    logger.error(f"Error occurred while processing project '{proj_name}': {e}")
    except Exception as e:
        logger.error(f"Error occurred while processing input file '{input_file}': {e}")
        return


if __name__ == "__main__":
    cwd = os.getcwd()
    clean_local_clone_path(os.path.join(cwd, 'clone'))
    print("Enter Username and Password")
    username = input("Please enter your username: ")
    password = input("Please enter your password: ")
    main(username, password)
